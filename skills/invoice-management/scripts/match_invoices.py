#!/usr/bin/env python3
"""
match_invoices.py — 凑票金额组合算法

读取发票统计表，用贪心算法选取组合覆盖目标金额，
创建日期文件夹并移入选中发票，更新统计表状态。

用法：
  python match_invoices.py --target 2000 --xlsx "发票统计表.xlsx" --unused-dir "未使用/" --work-dir "报销发票/" [--note "案件差旅费"]
"""

import argparse
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] 缺少 openpyxl，请安装: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def load_available_invoices(xlsx_path, unused_dir, filter_category=None):
    """
    从统计表读取未使用发票。
    返回 list of dict: {row, filename, category, amount, orig_name}
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    invoices = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[1]:  # 空行
            continue
        status = str(row[6]).strip() if row[6] else ""
        if status != "未使用":
            continue

        category = str(row[2]).strip() if row[2] else ""
        amount = float(row[3]) if row[3] else 0.0
        filename = str(row[1]).strip()

        if filter_category and category != filter_category:
            continue

        filepath = Path(unused_dir) / filename
        if not filepath.exists():
            print(f"  ⚠ 文件不存在，跳过: {filename}")
            continue

        invoices.append({
            "row": row_idx,
            "filename": filename,
            "category": category,
            "amount": amount,
            "filepath": str(filepath),
        })

    wb.close()
    return invoices


def greedy_match(invoices, target):
    """
    贪心算法选取组合。
    按金额降序，从大到小累加直到 >= target。
    返回 (selected, total, remainder)
    """
    sorted_inv = sorted(invoices, key=lambda x: x["amount"], reverse=True)
    selected = []
    total = 0.0

    for inv in sorted_inv:
        if total >= target:
            break
        selected.append(inv)
        total += inv["amount"]

    remainder = total - target
    return selected, round(total, 2), round(remainder, 2)


def create_folder_and_move(selected, work_dir, note=""):
    """
    创建 {日期}+{金额}/ 文件夹，移入选中发票。
    返回文件夹路径。
    """
    date_str = datetime.now().strftime("%Y-%m-%d")
    total = round(sum(inv["amount"] for inv in selected), 2)
    folder_name = f"{date_str}+{total:.2f}"
    folder_path = Path(work_dir) / folder_name
    folder_path.mkdir(parents=True, exist_ok=True)

    for inv in selected:
        src = inv["filepath"]
        dst = folder_path / inv["filename"]
        shutil.move(src, dst)

    return str(folder_path)


def update_xlsx(xlsx_path, selected, note=""):
    """更新统计表：状态→已使用，日期→当日，备注→note"""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    today = datetime.now().strftime("%Y-%m-%d")

    for inv in selected:
        row = inv["row"]
        ws.cell(row=row, column=7, value="已使用")
        ws.cell(row=row, column=8, value=today)
        ws.cell(row=row, column=9, value=note)

    wb.save(xlsx_path)


def merge_pdfs(selected, folder):
    """合并选中发票为一个PDF"""
    try:
        from pypdf import PdfWriter, PdfReader
    except ImportError:
        print("  ⚠ 缺少 pypdf，跳过合并 (pip install pypdf)")
        return None

    # 文件已移入 folder，用 folder + filename 构建路径
    pdfs = sorted([Path(folder) / inv["filename"] for inv in selected if (Path(folder) / inv["filename"]).suffix == '.pdf' and (Path(folder) / inv["filename"]).exists()])
    if len(pdfs) < 2:
        return None

    writer = PdfWriter()
    for pdf in pdfs:
        reader = PdfReader(pdf)
        for page in reader.pages:
            writer.add_page(page)

    merged = Path(folder) / "发票合并.pdf"
    writer.write(merged)
    return str(merged)


def fill_expense_form(selected, folder, work_dir):
    """填写费用报销单模板并保存到凑票文件夹"""
    from collections import defaultdict
    from openpyxl import load_workbook

    # 优先用skill内置模板，回退到工作目录
    skill_template = Path(__file__).parent.parent / "templates" / "费用报销单模板.xlsx"
    work_template = Path(work_dir) / "费用报销单.xlsx"
    template = skill_template if skill_template.exists() else work_template
    if not template.exists():
        print("  ⚠ 费用报销单模板不存在，跳过")
        return None

    # 按类型汇总金额
    cats = defaultdict(float)
    for inv in selected:
        cats[inv["category"]] += inv["amount"]

    # 按金额降序排列
    items = sorted(cats.items(), key=lambda x: x[1], reverse=True)

    wb = load_workbook(template)
    ws = wb["纵向"]

    for i, (cat, amt) in enumerate(items):
        if i >= 7:  # A4-A10 共7行
            break
        ws.cell(row=4 + i, column=1, value=cat)
        ws.cell(row=4 + i, column=2, value=round(amt, 2))

    dest = Path(folder) / "费用报销单.xlsx"
    wb.save(dest)
    return str(dest)


def main():
    parser = argparse.ArgumentParser(description="凑票金额组合")
    parser.add_argument("--target", type=float, required=True, help="目标金额")
    parser.add_argument("--xlsx", required=True, help="发票统计表.xlsx 路径")
    parser.add_argument("--unused-dir", required=True, help="未使用 目录")
    parser.add_argument("--work-dir", required=True, help="报销发票根目录")
    parser.add_argument("--note", default="", help="使用备注")
    parser.add_argument("--category", default=None, help='限定用途类型(如"交通费")')
    args = parser.parse_args()

    # 1. 加载可用发票
    print(f"🔍 读取可用发票 ...")
    invoices = load_available_invoices(args.xlsx, args.unused_dir, args.category)

    if not invoices:
        print("❌ 没有可用的未使用发票")
        sys.exit(1)

    total_available = round(sum(inv["amount"] for inv in invoices), 2)
    print(f"   可用 {len(invoices)} 张，合计 ¥{total_available:.2f}")

    if total_available < args.target:
        print(f"⚠ 库存不足：需要 ¥{args.target:.2f}，可用 ¥{total_available:.2f}，差额 ¥{total_available - args.target:.2f}")
        # 不退出，仍然用全部
        selected = invoices
        total = total_available
    else:
        selected, total, remainder = greedy_match(invoices, args.target)

        # 允许 ±15% 偏差
        deviation = abs(total - args.target) / args.target
        if deviation > 0.15 and total > args.target:
            print(f"⚠ 最佳组合 ¥{total:.2f} 偏离目标 {deviation:.1%}（允许最大 15%）")

    # 2. 创建文件夹并移入
    folder = create_folder_and_move(selected, args.work_dir, args.note)

    # 3. 更新统计表
    update_xlsx(args.xlsx, selected, args.note)

    # 4. 填写费用报销单
    expense_form = fill_expense_form(selected, folder, args.work_dir)

    # 5. 合并PDF
    merged_pdf = merge_pdfs(selected, folder)

    # 6. 汇报
    print(f"\n{'='*50}")
    print(f"📋 凑票完成 — 目标 ¥{args.target:.2f}")
    print(f"{'='*50}")
    print(f"{'用途':<10} {'金额':>10}")
    print(f"{'-'*22}")
    for inv in selected:
        print(f"{inv['category']:<10} {inv['amount']:>10.2f}")
    print(f"{'-'*22}")
    print(f"{'合计':<10} {total:>10.2f}")
    print(f"\n📁 已移入 {folder}/")
    if expense_form:
        print(f"📋 费用报销单: {expense_form}")
    if merged_pdf:
        print(f"📎 合并PDF: {merged_pdf}")
    print(f"📊 统计表已更新")
    print(f"💰 剩余可用 ¥{total_available - total:.2f}")


if __name__ == "__main__":
    main()
