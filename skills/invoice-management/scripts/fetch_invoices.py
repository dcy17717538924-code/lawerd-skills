#!/usr/bin/env python3
"""
fetch_invoices.py — 163邮箱发票下载+解析+归档

链路：POP3 → 本地去重 → 搜索发票邮件 → 下载PDF附件 → pdfplumber解析 → 重命名 → 写xlsx

用法：
  python fetch_invoices.py --email "xx@163.com" --password "授权码" [--days 30] [--save-dir "未使用/"]

去重机制：
  在 {save_dir}/.fetched_ids.json 记录已处理的邮件ID（基于 Date+Subject 哈希），
  下次运行时自动跳过，实现"只处理未读"的效果。
"""

import argparse
import email as em
import email.header
import hashlib
import json
import os
import poplib
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print("[ERROR] 缺少 pdfplumber", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("[ERROR] 缺少 openpyxl", file=sys.stderr)
    sys.exit(1)


CATEGORY_RULES = [
    (["交通卡充值", "客运服务费", "滴滴出行", "滴滴", "网约车", "出租车",
      "桔子出行", "雷利出行", "强生致行", "享道出行", "T3出行", "及时用车",
      "曹操出行", "喜行约车", "风韵出行", "聚的出租车", "高德打车",
      "机票", "携程", "航空", "飞行", "东潮出行", "有序出行", "阳光出行",
      "妥妥E行", "携华出行", "胖哒出行", "麦卡出行", "哈啰"], "交通费"),
    (["餐饮服务", "餐饮", "食品", "汉堡王", "韩宴", "小吃", "火锅"], "餐饮费"),
    (["通信服务费", "电信服务", "联通", "中国联通", "移动通信"], "通讯费"),
    (["住宿", "酒店", "旅馆", "大酒店"], "住宿费"),
]

# 只下载 PDF 发票，跳过 OFD/XML/行程单
INVOICE_ATTACHMENT_EXTS = {".pdf"}
# 文件名黑名单：行程报销单不是发票
SKIP_FILENAME_KEYWORDS = ["行程报销单", "行程单"]


def decode_header(raw):
    parts = em.header.decode_header(raw)
    result = []
    for part, charset in parts:
        if isinstance(part, bytes):
            result.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            result.append(part)
    return "".join(result)


def extract_amount(text):
    patterns = [
        r"价税合计[：:]?\s*[¥￥]\s*([\d,]+\.\d{2})",
        r"合计金额[：:]?\s*[¥￥]\s*([\d,]+\.\d{2})",
        r"合计[：:]?\s*[¥￥]\s*([\d,]+\.\d{2})",
        r"金额[：:]?\s*[¥￥]\s*([\d,]+\.\d{2})",
        r"发票金额[：:]?\s*[¥￥]?\s*([\d,]+\.?\d{0,2})",
    ]
    amounts = []
    for pat in patterns:
        for m in re.findall(pat, text):
            try:
                amounts.append(float(m.replace(",", "")))
            except ValueError:
                pass
    if not amounts:
        for m in re.findall(r"[¥￥]\s*([\d,]+\.\d{2})", text):
            try:
                amounts.append(float(m.replace(",", "")))
            except ValueError:
                pass
    return round(max(amounts), 2) if amounts else None


def classify(text, filename=""):
    search_text = (text + " " + filename).lower()
    for keywords, category in CATEGORY_RULES:
        for kw in keywords:
            if kw in search_text:
                return category
    return "其他"


def extract_text_from_pdf(filepath):
    try:
        with pdfplumber.open(filepath) as pdf:
            text = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
            return text
    except Exception as e:
        print(f"  ⚠ pdfplumber: {e}")
        return ""


def load_fetched_ids(track_file):
    if track_file.exists():
        return set(json.loads(track_file.read_text()))
    return set()


def save_fetched_ids(track_file, ids):
    track_file.write_text(json.dumps(list(ids)))


def append_xlsx(xlsx_path, records):
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
        max_seq = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and isinstance(row[0], (int, float)):
                max_seq = max(max_seq, int(row[0]))
        seq_start = max_seq + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "发票统计"
        headers = ["序号", "新文件名", "用途类型", "金额（元）", "发票详情", "原文件名", "状态", "使用日期", "使用备注"]
        hfont = Font(bold=True, size=11, color="FFFFFF")
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
        tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = tb
        for col, w in zip("ABCDEFGHI", [6,28,10,12,25,45,10,14,20]):
            ws.column_dimensions[col].width = w
        seq_start = 1

    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    for i, rec in enumerate(records):
        row = seq_start + i + 1
        ws.cell(row=row, column=1, value=seq_start + i).border = tb
        ws.cell(row=row, column=2, value=rec["new_name"]).border = tb
        ws.cell(row=row, column=3, value=rec["category"]).border = tb
        c = ws.cell(row=row, column=4, value=rec["amount"]); c.border = tb; c.number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=rec["detail"]).border = tb
        ws.cell(row=row, column=6, value=rec["orig_name"]).border = tb
        ws.cell(row=row, column=7, value="未使用").border = tb
        ws.cell(row=row, column=8).border = tb
        ws.cell(row=row, column=9).border = tb
    wb.save(xlsx_path)


def should_skip_filename(filename):
    for kw in SKIP_FILENAME_KEYWORDS:
        if kw in filename:
            return True
    return False


def process_attachments(msg, save_dir, sender):
    results = []
    subject = decode_header(msg["Subject"] or "")[:60]
    print(f"\n📩 {subject}...")

    for part in msg.walk():
        orig_filename = part.get_filename()
        if not orig_filename:
            continue
        disp = str(part.get("Content-Disposition", ""))
        if not orig_filename:
            continue
        orig_filename = decode_header(orig_filename)
        ext = Path(orig_filename).suffix.lower()

        # 只下载 PDF 发票
        if ext not in INVOICE_ATTACHMENT_EXTS:
            continue
        if should_skip_filename(orig_filename):
            print(f"  ⏭ 跳过非发票: {orig_filename}")
            continue

        dest_path = save_dir / orig_filename
        if dest_path.exists():
            print(f"  ⏭ 已存在: {orig_filename}")
            continue

        payload = part.get_payload(decode=True)
        with open(dest_path, "wb") as f:
            f.write(payload)
        print(f"  ✅ 下载: {orig_filename}")

        # 解析 PDF
        text = extract_text_from_pdf(str(dest_path))
        amount = None
        category = "其他"
        detail = ""

        if not text.strip():
            print(f"  ⚠ pdfplumber无文字,尝试markitdown...")
            try:
                import subprocess
                r = subprocess.run([
                    r"{{CODEX_PYTHON}}",
                    "-m", "markitdown", str(dest_path),
                ], capture_output=True, text=True, timeout=30)
                if r.returncode == 0:
                    text = r.stdout
            except Exception:
                pass

        if text.strip():
            amount = extract_amount(text)
            category = classify(text, orig_filename)
            detail = text[:80].replace("\n", " ").strip()
        else:
            print(f"  ⚠ 无法提取文字，按文件名分类")
            category = classify("", orig_filename)

        if amount:
            new_name = f"{category}{amount:.2f}{ext}"
        else:
            new_name = f"{category}_待确认_{orig_filename}"

        new_path = save_dir / new_name
        counter = 1
        while new_path.exists():
            new_name = f"{category}{amount:.2f}_{counter}{ext}" if amount else f"{category}_待确认_{counter}_{orig_filename}"
            new_path = save_dir / new_name
            counter += 1

        os.rename(dest_path, new_path)
        print(f"  📝 → {new_name}  ¥{amount or '?'}")

        results.append({
            "new_name": new_name, "category": category,
            "amount": amount or 0.00,
            "detail": detail or f"发件人: {sender[:40]}",
            "orig_name": orig_filename,
        })
    return results


def main():
    parser = argparse.ArgumentParser(description="163邮箱发票下载+解析+归档(POP3)")
    parser.add_argument("--email", required=True)
    parser.add_argument("--password", required=True)
    parser.add_argument("--save-dir", required=True)
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--search-keyword", default="发票|报销凭证|电子凭证|数电票|全电发票", help="多个关键词用|分隔")
    parser.add_argument("--days", type=int, default=90)
    args = parser.parse_args()

    save_dir = Path(args.save_dir)
    save_dir.mkdir(parents=True, exist_ok=True)

    # 加载去重记录
    track_file = save_dir / ".fetched_ids.json"
    fetched = load_fetched_ids(track_file)
    print(f"📋 已处理 {len(fetched)} 封邮件（本地去重）")

    # POP3
    print("📧 POP3 (pop.163.com:995)...")
    mail = poplib.POP3_SSL("pop.163.com", 995)
    mail.user(args.email)
    mail.pass_(args.password)
    count, _ = mail.stat()
    print(f"   📬 {count} 封邮件")

    cutoff = datetime.now() - timedelta(days=args.days)
    new_records = []
    matched = 0
    skipped_dedup = 0
    new_fetched = set()

    from email import policy
    from email.utils import parsedate_to_datetime

    for i in range(count, 0, -1):
        resp, lines, _ = mail.retr(i)
        raw = b"\r\n".join(lines)
        msg = em.message_from_bytes(raw, policy=policy.default)

        # 日期过滤
        date_str = msg["Date"]
        if date_str:
            try:
                msg_date = parsedate_to_datetime(date_str)
                if msg_date.replace(tzinfo=None) < cutoff:
                    break
            except Exception:
                pass

        # 主题过滤
        subject = decode_header(msg["Subject"] or "")
        keywords = args.search_keyword.split("|")
        if not any(kw in subject for kw in keywords):
            continue

        # 去重：基于 Date+Subject 哈希
        dedup_key = hashlib.md5(
            f"{msg['Date']}|{subject}".encode()
        ).hexdigest()
        if dedup_key in fetched:
            skipped_dedup += 1
            continue

        new_fetched.add(dedup_key)
        matched += 1
        sender = decode_header(msg["From"] or "")
        new = process_attachments(msg, save_dir, sender)
        new_records.extend(new)

        if matched >= 100:
            break

    mail.quit()

    # 保存去重记录
    fetched.update(new_fetched)
    save_fetched_ids(track_file, fetched)

    # 写入 xlsx
    if new_records:
        append_xlsx(args.xlsx, new_records)
        print(f"\n📊 统计表已更新: {args.xlsx}")

    # 汇报
    print(f"\n{'='*50}")
    print(f"📧 邮箱发票下载完成（POP3）")
    print(f"{'='*50}")
    print(f"  匹配邮件: {matched} 封（跳过 {skipped_dedup} 封已处理）")
    print(f"  新增归档: {len(new_records)} 份")
    print(f"  去重记录: {len(fetched)} 封")
    if new_records:
        print(f"\n  新增清单:")
        total_amt = sum(r["amount"] for r in new_records)
        for r in new_records:
            print(f"    {r['new_name']:<40} ¥{r['amount']:.2f}")
        print(f"  💰 合计: ¥{total_amt:.2f}")


if __name__ == "__main__":
    main()
