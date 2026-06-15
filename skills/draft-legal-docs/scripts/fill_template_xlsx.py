#!/usr/bin/env python3
"""证据目录 XLSX 模板填充脚本 — openpyxl 文本替换 + 动态行管理

用法:
  python fill_template_xlsx.py 证据目录.template.xlsx 输出.xlsx --data data.json
  python fill_template_xlsx.py 证据目录.template.xlsx 输出.xlsx --提交人 "原告张三" --日期 "2026-06-15"

JSON 数据格式:
{
  "提交人": "原告：张三",
  "目录制作日期": "2026-06-15",
  "证据分组": [
    {
      "模式": "小类型",
      "项目": ["借款合同", "补充协议"],
      "证明目的": "证明双方存在借贷合意",
      "页码": "1-7"
    },
    {
      "模式": "大类型",
      "项目": ["银行转账凭证"],
      "证明目的": "证明原告已履行出借义务",
      "页码": "8-11"
    }
  ]
}

「模式」:
  - "小类型": 多子项，中文编号列纵向合并，B列数字子编号(1,2,3…)
  - "大类型": 单行，中文编号独占A+B合并格
省略模式时默认 "小类型"。
"""

import openpyxl, json, sys, argparse, shutil
from pathlib import Path
from openpyxl.styles import Font, Alignment, Border, Side

# ── 中文数字（支持到九十九） ──────────────────────────────────
_CN = '零一二三四五六七八九十'
def _cn_num(n: int) -> str:
    if n <= 10:
        return _CN[n]
    if n < 20:
        return '十' + _CN[n - 10]
    tens = n // 10
    ones = n % 10
    s = _CN[tens] + '十'
    if ones:
        s += _CN[ones]
    return s


# ── 默认样式（从模板提取） ──────────────────────────────────
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
FONT_DEFAULT = Font(name='宋体', size=12)
FONT_TITLE  = Font(name='宋体', size=16)
FONT_DATE   = Font(name='宋体', size=11)
ALIGN_CENTER       = Alignment(horizontal='center', vertical='center')
ALIGN_CENTER_WRAP  = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_WRAP    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_RIGHT        = Alignment(horizontal='right',  vertical='center')
NO_BORDER = Border()


def _apply_cell(cell, value=None, font=FONT_DEFAULT, alignment=ALIGN_CENTER,
                border=THIN_BORDER, number_format='General'):
    cell.value = value
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    cell.number_format = number_format


# ═══════════════════════════════════════════════════════════
#  核心填充逻辑
# ═══════════════════════════════════════════════════════════

def fill_xlsx(template_path: str, output_path: str,
              submitter: str = None, date_str: str = None,
              groups: list = None):
    """填充证据目录 XLSX 模板。"""
    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # ── 1. 替换标题区占位符 ──
    if submitter:
        ws['A2'] = f'提交人 ： {submitter}'
        _apply_cell(ws['A2'], ws['A2'].value, font=FONT_DEFAULT,
                    alignment=ALIGN_LEFT_WRAP, border=NO_BORDER)

    # ── 2. 解除原数据区合并 ──
    _unmerge_data_rows(ws)
    DATA_START = 4
    DATA_END   = 12

    # ── 3. 按 groups 重建数据行 ──
    if groups:
        current_row = DATA_START
        for i, g in enumerate(groups, 1):
            编号 = _cn_num(i)
            mode = g.get('模式', '小类型')
            items = g.get('项目', [])
            if isinstance(items, str):
                items = [items]
            items = [it['名称'] if isinstance(it, dict) else str(it) for it in items]
            if not items:
                items = ['']

            证明目的 = g.get('证明目的', '')
            页码     = g.get('页码', '')

            if mode == '大类型':
                current_row = _write_big_group(ws, current_row, 编号, items[0],
                                               证明目的, 页码)
            else:
                current_row = _write_small_group(ws, current_row, 编号, items,
                                                 证明目的, 页码)
    else:
        current_row = DATA_START

    # ── 4. 清理多余旧行（含旧日期行 R13） ──
    _clear_remaining_rows(ws, current_row, max(DATA_END, 13))

    # ── 5. 日期落款（数据区后空一行） ──
    date_row = current_row + 1
    _apply_cell(ws.cell(row=date_row, column=4),
                value=f'目录制作日期：{date_str or "YY-MM-DD"}',
                font=FONT_DATE, alignment=ALIGN_RIGHT, border=NO_BORDER)

    # ── 6. 保存 ──
    wb.save(output_path)
    return output_path


# ── 辅助 ──

def _unmerge_data_rows(ws):
    to_remove = []
    for mc in ws.merged_cells.ranges:
        if mc.min_row >= 4:
            to_remove.append(str(mc))
    for r in to_remove:
        ws.unmerge_cells(r)


def _write_big_group(ws, row, 编号: str, name: str, 证明目的: str, 页码: str) -> int:
    """A+B 合并写中文编号，单行。返回下一可用行。"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    _apply_cell(ws.cell(row=row, column=1), value=编号, alignment=ALIGN_CENTER_WRAP)
    # B列在合并区内，不单独写值（MergedCell 只读）
    _apply_cell(ws.cell(row=row, column=3), value=name)
    _apply_cell(ws.cell(row=row, column=4), value=证明目的, alignment=ALIGN_CENTER_WRAP)
    _apply_cell(ws.cell(row=row, column=5), value=页码, number_format='@')
    ws.row_dimensions[row].height = 38
    return row + 1


def _write_small_group(ws, row, 编号: str, items: list,
                       证明目的: str, 页码: str) -> int:
    """A/D/E 纵向合并，B 数字子编号，C 子项名称。返回下一可用行。"""
    n = len(items)
    end_row = row + n - 1

    if n > 1:
        for col in (1, 4, 5):
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=end_row, end_column=col)

    _apply_cell(ws.cell(row=row, column=1), value=编号, alignment=ALIGN_CENTER_WRAP)
    _apply_cell(ws.cell(row=row, column=4), value=证明目的, alignment=ALIGN_CENTER_WRAP)
    _apply_cell(ws.cell(row=row, column=5), value=页码, number_format='@')

    for j, item_name in enumerate(items):
        r = row + j
        _apply_cell(ws.cell(row=r, column=2), value=j + 1, alignment=ALIGN_CENTER_WRAP)
        _apply_cell(ws.cell(row=r, column=3), value=item_name)
        ws.row_dimensions[r].height = 38

    return end_row + 1


def _clear_remaining_rows(ws, from_row, up_to_row):
    for r in range(from_row, up_to_row + 1):
        for c in range(1, 6):
            try:
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.border = NO_BORDER
                cell.font = FONT_DEFAULT
            except AttributeError:
                pass  # MergedCell 跳过


# ═══════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='证据目录 XLSX 模板填充')
    parser.add_argument('template', help='模板路径（.xlsx）')
    parser.add_argument('output',   help='输出路径（.xlsx）')
    parser.add_argument('--data',   help='JSON 数据文件')
    parser.add_argument('--提交人', help='证据提交方')
    parser.add_argument('--日期',   help='目录制作日期（YYYY-MM-DD）')
    args = parser.parse_args()

    submitter = args.提交人
    date_str  = args.日期
    groups    = None

    if args.data:
        with open(args.data, 'r', encoding='utf-8') as f:
            d = json.load(f)
        submitter = d.get('提交人', submitter)
        date_str  = d.get('目录制作日期', date_str)
        groups    = d.get('证据分组')

    if not groups:
        print('WARNING: no evidence groups provided, only title placeholders will be replaced.')

    out = fill_xlsx(args.template, args.output,
                    submitter=submitter, date_str=date_str, groups=groups)
    print(f'Done: {out}')


if __name__ == '__main__':
    main()
